import sys
sys.path.append("/Users/mikitosaarna/Downloads")
import openpyxl
import os

class ConvertSheet:
    def __init__(self, csv_file_path):

      column_labels = ["First name", "Last name", "SSN", "Date of Birth", "Cell phone", "Email", "Property Address", "City", "State", "Zip"]

      date_row = {
        "name": 13,
        "ssn": 14,
        "dob": 15,
        "cell": 17,
        "email": 19,
        "address": 25,
        "city/state/zip": 26
      }

      new_file = openpyxl.Workbook()
      new_spreadsheet = new_file.active
      new_spreadsheet.append(column_labels)

      file = openpyxl.load_workbook(csv_file_path)
      spreadsheet = file.active
      first_row = spreadsheet[1]

      for index, cell in enumerate(first_row):
        if (cell.value is not None and "LOAN SUMMARYLoan Number:" in cell.value):
          search_column_index = index + 2

          name = spreadsheet.cell(row=date_row["name"], column=search_column_index)
          name_array = name.value.split(" ")
          first_name = name_array[0]
          last_name = name_array[len(name_array) - 1]
          ssn = spreadsheet.cell(row=date_row["ssn"], column=search_column_index).value

          dob_unformatted = spreadsheet.cell(row=date_row["dob"], column=search_column_index).value
          dob_formatted = dob_unformatted.strftime('%m/%d/%Y')

          cell = spreadsheet.cell(row=date_row["cell"], column=search_column_index).value
          email = spreadsheet.cell(row=date_row["email"], column=search_column_index).value
          address = spreadsheet.cell(row=date_row["address"], column=search_column_index).value

          city_state_zip = spreadsheet.cell(row=date_row["city/state/zip"], column=search_column_index)
          city_state_zip_array = city_state_zip.value.split(" ")
          city = city_state_zip_array[0][:-1]
          state = city_state_zip_array[1]
          zip = city_state_zip_array[2]

          data_row = [
            first_name,
            last_name,
            ssn,
            dob_formatted,
            cell,
            email,
            address,
            city,
            state,
            zip
          ]

          new_spreadsheet.append(data_row)

      home_directory = os.path.expanduser('~')
      save_file_path = os.path.join(home_directory, 'Downloads', 'new_file_1.xlsx')
      new_file.save(save_file_path)

if __name__ == "__main__":
  # csv_file_path = '../../../Downloads/example_sheet.xlsx' # "~/Downloads/example_sheet - Sheet1.csv"
  csv_file_path = '../../../Downloads/mortgage closings 1.2 (1) (1).xlsx'
  ConvertSheet(csv_file_path)